from __future__ import annotations

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Literal, cast

from src.utils.json_types import JsonObject

_UNNAMED_RE = re.compile(r"^Unnamed:\s*\d+$")


def _jsonable_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    item = getattr(value, "item", None)
    if callable(item):
        return _jsonable_value(item())
    return str(value)


def _normalize_column_names(raw_names: list[object]) -> list[str]:
    normalized: list[str] = []
    seen: dict[str, int] = {}
    for idx, raw in enumerate(raw_names, start=1):
        if raw is None or isinstance(raw, int):
            candidate = ""
        else:
            candidate = str(raw).replace("\n", " ").strip()
        if candidate == "" or candidate.lower() == "nan" or _UNNAMED_RE.match(candidate):
            candidate = f"col_{idx}"
        base = candidate
        count = seen.get(base, 0) + 1
        seen[base] = count
        normalized.append(base if count == 1 else f"{base}_{count}")
    return normalized


def _apply_normalized_columns(*, df: Any) -> Any:
    cols = list(getattr(df, "columns", []))
    df.columns = _normalize_column_names(cols)
    return df


def _infer_row_count_csv(*, path: Path, header_row: bool) -> int | None:
    try:
        with path.open("rb") as handle:
            total = sum(1 for _ in handle)
    except OSError:
        return None
    if total <= 0:
        return 0
    return max(total - (1 if header_row else 0), 0)


def _csv_preview(*, path: Path, rows: int, columns: int, header_row: bool) -> JsonObject:
    import pandas as pd

    header: int | None = 0 if header_row else None
    df = pd.read_csv(path, nrows=rows, header=header)
    df = _apply_normalized_columns(df=df)
    column_count: int | None
    try:
        header_df = pd.read_csv(path, nrows=0, header=header)
        column_count = int(header_df.shape[1])
    except (ValueError, OSError):
        column_count = int(df.shape[1])
    return cast(
        JsonObject,
        {
            "row_count": _infer_row_count_csv(path=path, header_row=header_row),
            "column_count": column_count,
            "columns": _infer_columns(df=df, columns=columns),
            "sample_rows": _sample_rows(df=df, columns=columns),
        },
    )


def excel_sheet_names(*, path: Path) -> list[str]:
    if path.suffix.lower() == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path))
        return [name for name in book.sheet_names() if isinstance(name, str) and name.strip()]

    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        visible: list[str] = []
        for ws in workbook.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            name = getattr(ws, "title", None)
            if isinstance(name, str) and name.strip():
                visible.append(name)
        return visible
    finally:
        workbook.close()


def _excel_sheet_meta(
    *, path: Path, sheet_name: str | None
) -> tuple[list[str], str | None, int | None, int | None, list[object], list[object]]:
    names = excel_sheet_names(path=path)
    selected = sheet_name if sheet_name in names else (names[0] if len(names) > 0 else None)
    if selected is None:
        return names, None, None, None, [], []

    if path.suffix.lower() == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_name(selected)
        first = sheet.row_values(0) if sheet.nrows >= 1 else []
        second = sheet.row_values(1) if sheet.nrows >= 2 else []
        return names, selected, sheet.nrows, sheet.ncols, list(first), list(second)

    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook[selected]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        second_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
        return (
            names,
            selected,
            int(ws.max_row),
            int(ws.max_column),
            list(first_row),
            list(second_row),
        )
    finally:
        workbook.close()


_NUMERIC_RE = re.compile(r"^[-+]?\d+(?:\.\d+)?$")


def _looks_numeric(value: str) -> bool:
    if value.strip() == "":
        return False
    return _NUMERIC_RE.match(value.strip()) is not None


def _infer_header_row_from_rows(*, first_row: list[object], second_row: list[object]) -> bool:
    first = ["" if v is None else str(v).strip() for v in first_row]
    second = ["" if v is None else str(v).strip() for v in second_row]

    non_empty_first = [v for v in first if v != ""]
    if len(non_empty_first) == 0:
        return False

    string_like = sum(1 for v in non_empty_first if not _looks_numeric(v)) / len(non_empty_first)
    unique_ratio = len(set(non_empty_first)) / len(non_empty_first)
    if string_like >= 0.7 and unique_ratio >= 0.7:
        return True

    non_empty_second = [v for v in second if v != ""]
    numeric_ratio_second = (
        sum(1 for v in non_empty_second if _looks_numeric(v)) / len(non_empty_second)
        if len(non_empty_second) > 0
        else 0.0
    )
    return string_like >= 0.5 and numeric_ratio_second >= 0.5


def _excel_preview(
    *,
    path: Path,
    rows: int,
    columns: int,
    sheet_name: str | None,
    header_row: bool | None,
) -> JsonObject:
    import pandas as pd

    sheet_names, selected, raw_rows, raw_cols, first_row, second_row = _excel_sheet_meta(
        path=path, sheet_name=sheet_name
    )
    effective_header = header_row if header_row is not None else _infer_header_row_from_rows(
        first_row=first_row, second_row=second_row
    )
    header: int | None = 0 if effective_header else None
    engine: Literal["xlrd", "openpyxl"] = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    engine_kwargs: dict[str, object] | None = None
    if engine == "openpyxl":
        engine_kwargs = {"data_only": False}
    df = pd.read_excel(
        path,
        nrows=rows,
        sheet_name=selected,
        engine=engine,
        header=header,
        engine_kwargs=engine_kwargs,
    )
    df = _apply_normalized_columns(df=df)
    row_count = None if raw_rows is None else max(int(raw_rows) - (1 if effective_header else 0), 0)
    return cast(
        JsonObject,
        {
            "row_count": row_count,
            "column_count": None if raw_cols is None else int(raw_cols),
            "columns": _infer_columns(df=df, columns=columns),
            "sample_rows": _sample_rows(df=df, columns=columns),
            "sheet_names": sheet_names,
            "selected_sheet": selected,
            "header_row": effective_header,
        },
    )


def _dta_preview(*, path: Path, rows: int, columns: int) -> JsonObject:
    import pandas as pd
    from pandas.io.stata import StataReader

    try:
        with StataReader(path) as reader:
            row_count = int(cast(Any, reader).nobs)
            column_count = int(cast(Any, reader).nvar)
    except (OSError, ValueError):
        row_count = None
        column_count = None

    it = pd.read_stata(path, chunksize=rows)
    try:
        df = next(it)
    except StopIteration:
        df = pd.DataFrame()
    df = _apply_normalized_columns(df=df)
    return cast(
        JsonObject,
        {
            "row_count": row_count,
            "column_count": column_count,
            "columns": _infer_columns(df=df, columns=columns),
            "sample_rows": _sample_rows(df=df, columns=columns),
        },
    )


def _infer_columns(*, df: Any, columns: int) -> list[dict[str, str]]:
    import pandas as pd

    types = pd.api.types
    limited = df.iloc[:, :columns]
    columns_preview: list[dict[str, str]] = []
    for col in limited.columns:
        dtype = limited[col].dtype
        inferred = "unknown"
        if types.is_bool_dtype(dtype):
            inferred = "boolean"
        elif types.is_integer_dtype(dtype):
            inferred = "integer"
        elif types.is_float_dtype(dtype):
            inferred = "number"
        elif types.is_datetime64_any_dtype(dtype):
            inferred = "datetime"
        elif types.is_string_dtype(dtype) or types.is_object_dtype(dtype):
            inferred = "string"
        columns_preview.append({"name": str(col), "inferred_type": inferred})
    return columns_preview


def _sample_rows(
    *,
    df: Any,
    columns: int,
) -> list[dict[str, str | int | float | bool | None]]:
    limited = df.iloc[:, :columns]
    raw_rows = limited.to_dict(orient="records")
    sample_rows: list[dict[str, str | int | float | bool | None]] = []
    for row in cast(list[dict[str, Any]], raw_rows):
        sample_rows.append({str(k): _jsonable_value(v) for k, v in row.items()})
    return sample_rows


def dataset_preview(*, path: Path, fmt: str, rows: int, columns: int) -> JsonObject:
    return dataset_preview_with_options(path=path, fmt=fmt, rows=rows, columns=columns)


def dataset_preview_with_options(
    *,
    path: Path,
    fmt: str,
    rows: int,
    columns: int,
    sheet_name: str | None = None,
    header_row: bool | None = None,
) -> JsonObject:
    if fmt == "csv":
        csv_header_row = header_row is not False
        return _csv_preview(path=path, rows=rows, columns=columns, header_row=csv_header_row)
    if fmt == "excel":
        return _excel_preview(
            path=path,
            rows=rows,
            columns=columns,
            sheet_name=sheet_name,
            header_row=header_row,
        )
    if fmt == "dta":
        return _dta_preview(path=path, rows=rows, columns=columns)
    raise ValueError(f"unsupported format: {fmt}")
